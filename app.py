import customtkinter as ctk
from tkinter import filedialog
from threading import Thread
import pandas as pd
from faster_whisper import WhisperModel
import os
from openpyxl import load_workbook
from openpyxl.styles import Alignment

import sys

# Updated the UI to make it more modern and visually appealing
class TranscriberApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("Wazza Studios - AI Video Transcriber Pro")
        self.geometry("700x600")
        self.configure(fg_color="#2b2b2b")  # Dark background
        
        # Determine base path for resources
        if getattr(sys, 'frozen', False):
            # Running as compiled executable
            base_path = sys._MEIPASS
        else:
            # Running as script
            base_path = os.path.dirname(os.path.abspath(__file__))

        # Ensure ffmpeg in bundle or current directory is found
        ffmpeg_name = "ffmpeg.exe" if os.name == 'nt' else "ffmpeg"
        ffmpeg_path = os.path.join(base_path, ffmpeg_name)
        
        if os.path.exists(ffmpeg_path):
            os.environ["PATH"] += os.pathsep + os.path.dirname(ffmpeg_path)
            print(f"Set FFMPEG path to: {ffmpeg_path}")
            self.ffmpeg_path = ffmpeg_path
        else:
            # Fallback to system path or local dir if not bundled/found
            if os.path.exists(ffmpeg_name):
                 self.ffmpeg_path = os.path.abspath(ffmpeg_name)
            else:
                 self.ffmpeg_path = ffmpeg_name # Rely on system PATH



        # Control Variables
        self.is_running = False
        self.cancel_requested = False

        # Header
        self.header = ctk.CTkLabel(self, text="Wazza Studios - AI Video Transcriber Pro", font=("Arial", 24, "bold"), text_color="#ffffff")
        self.header.pack(pady=20)

        # Select Video Button
        self.btn_select = ctk.CTkButton(self, text="Select a Video", command=self.select_file, fg_color="#1f6aa5", hover_color="#144e78", text_color="#ffffff")
        self.btn_select.pack(pady=10)

        # Cancel Button (Hidden by default)
        self.btn_cancel = ctk.CTkButton(self, text="Cancel Task", fg_color="#d9534f", hover_color="#c9302c", text_color="#ffffff", command=self.request_cancel)

        # Progress Bar and Label
        self.progress_label = ctk.CTkLabel(self, text="Progress: 0%", font=("Arial", 14), text_color="#d1d1d1")
        self.progress_bar = ctk.CTkProgressBar(self, width=500, progress_color="#1f6aa5", fg_color="#3a3a3a")
        self.progress_bar.set(0)

        # Status Log
        self.status_log = ctk.CTkTextbox(self, width=600, height=200, fg_color="#3a3a3a", text_color="#ffffff", font=("Consolas", 12))
        self.status_log.pack(pady=20)

        # Footer
        self.footer = ctk.CTkLabel(self, text="Powered by Wazza Studios", font=("Arial", 12, "italic"), text_color="#d1d1d1")
        self.footer.pack(side="bottom", pady=10)

    def log(self, message):
        self.status_log.insert("end", message + "\n")
        self.status_log.see("end")

    def request_cancel(self):
        if self.is_running:
            self.cancel_requested = True
            self.log("!!! Cancellation requested. Stopping...")
            self.btn_cancel.configure(state="disabled", text="Stopping...")

    def select_file(self):
        file_path = filedialog.askopenfilename(filetypes=[("Video files", "*.mp4 *.mkv *.mov *.avi")])
        if file_path:
            # UI Preparation
            self.status_log.delete("1.0", "end") 
            self.progress_bar.set(0)
            self.progress_label.configure(text="Progress: 0%")
            
            self.btn_select.configure(state="disabled")
            self.btn_cancel.pack(pady=5)
            self.progress_bar.pack(pady=10)
            self.progress_label.pack(pady=5)
            
            self.is_running = True
            self.cancel_requested = False
            Thread(target=self.process_video, args=(file_path,), daemon=True).start()

    def process_video(self, video_path):
        try:
            self.log("Loading AI Model...")
            model = WhisperModel("small", device="cpu", compute_type="int8", cpu_threads=4)
            
            self.log("Converting video audio for processing...")
            
            # Use a temp wav file
            temp_wav = "temp_audio.wav"
            
            # Check for local ffmpeg
            ffmpeg_exe = self.ffmpeg_path

            # Convert to wav
            import subprocess
            # -y overwrites, -i input, -vn no video, -acodec pcm_s16le, -ar 16000 (whisper optimal), -ac 1 mono
            cmd = [
                ffmpeg_exe, "-y", 
                "-i", video_path, 
                "-vn", 
                "-acodec", "pcm_s16le", 
                "-ar", "16000", 
                "-ac", "1", 
                temp_wav
            ]
            
            startupinfo = None
            if os.name == 'nt':
                startupinfo = subprocess.STARTUPINFO()
                startupinfo.dwFlags |= subprocess.STARTF_USESHOWWINDOW
            
            try:
                subprocess.run(cmd, check=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE, startupinfo=startupinfo)
            except Exception as e:
                self.log(f"Error converting audio: {e}")
                self.log("Check if ffmpeg is installed/accessible.")
                return

            self.log("Transcribing...")
            self.log("This may take some time depending on video length...")
            
            segments_generator, info = model.transcribe(
                temp_wav, 
                beam_size=5, 
                word_timestamps=True,  # Enable word timestamps for precise cutting
                vad_filter=True
            )
            
            total_duration = info.duration
            rows = []
            
            # Excel Time Formatters
            fmt_excel = lambda s: f"{int(s//3600):02}:{int((s%3600)//60):02}:{int(s%60):02}"
            
            # Explicitly loop to track progress
            buffer_start = None
            buffer_words = []
            
            for segment in segments_generator:
                if self.cancel_requested:
                    self.log("Task cancelled.")
                    break
                
                # Check directly on words if available
                if not segment.words:
                    continue

                for word in segment.words:
                    if buffer_start is None:
                        buffer_start = word.start
                    
                    buffer_words.append(word.word.strip())
                    current_end = word.end
                    
                    # Check if we reached the target duration (5 seconds)
                    if (current_end - buffer_start) >= 5.0:
                        text_content = " ".join(buffer_words)
                        start_sec = buffer_start
                        end_sec = current_end
                        duration_sec = end_sec - start_sec
                        
                        row = [
                            os.path.basename(video_path), 
                            "", "", "", 
                            fmt_excel(start_sec), 
                            fmt_excel(end_sec), 
                            f"00:00:{int(duration_sec):02}", 
                            "", "", "", 
                            text_content, 
                            "", "", ""
                        ]
                        rows.append(row)
                        
                        # Reset buffer
                        buffer_start = None
                        buffer_words = []

                # Update progress using the segment end
                if total_duration > 0:
                    progress = min(segment.end / total_duration, 0.99)
                    self.progress_bar.set(progress)
                    self.progress_label.configure(text=f"Progress: {int(progress * 100)}%")
            
            # Flush any remaining buffered words
            if buffer_words and not self.cancel_requested:
                 text_content = " ".join(buffer_words)
                 start_sec = buffer_start
                 # Use the last known current_end from the loop
                 end_sec = current_end 
                 duration_sec = end_sec - start_sec
                 
                 row = [
                    os.path.basename(video_path), 
                    "", "", "", 
                    fmt_excel(start_sec), 
                    fmt_excel(end_sec), 
                    f"00:00:{int(duration_sec):02}", 
                    "", "", "", 
                    text_content, 
                    "", "", ""
                ]
                 rows.append(row)
                
            if self.cancel_requested:
                # Cleanup
                if os.path.exists(temp_wav):
                    os.remove(temp_wav)
                return

            self.log("Saving Excel...")
            output_file = video_path.rsplit(".", 1)[0] + "_script.xlsx"
            self.save_to_template(rows, output_file)
            
            self.progress_bar.set(1.0)
            self.progress_label.configure(text="Progress: 100%")
            self.log(f"Success! Saved to: {output_file}")
            
            # Cleanup
            if os.path.exists(temp_wav):
                os.remove(temp_wav)
            
        except Exception as e:
            self.log(f"Error: {str(e)}")
            import traceback
            traceback.print_exc()
        finally:
            self.is_running = False
            self.btn_select.configure(state="normal")
            self.btn_cancel.pack_forget()
            self.btn_cancel.configure(state="normal", text="Cancel Task")

    def save_to_template(self, rows_data, output_path):
        headers = [
            "Video", "Background music", "Theme", "Guest", "Start time", 
            "End time", "Time length", "audio", "start time", "end time", 
            "Phrase", "Angle", "Comment", "Additional comment"
        ]
        
        df = pd.DataFrame(rows_data, columns=headers)
        df.to_excel(output_path, index=False, startrow=3)

        # Formatting
        wb = load_workbook(output_path)
        ws = wb.active
        column_widths = {'A': 25, 'E': 12, 'F': 12, 'K': 65}
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        for row in ws.iter_rows(min_row=4):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
        
        wb.save(output_path)

if __name__ == "__main__":
    app = TranscriberApp()
    app.mainloop()